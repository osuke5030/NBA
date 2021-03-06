import requests
from bs4 import BeautifulSoup
import os
import json
import sys
sys.path.append("/Users/kuramochiosuke/.pyenv/versions/3.10.4/Library/Frameworks/Python.framework/Versions/3.10/lib/python3.10/site-packages")
from requests_oauthlib import OAuth1Session
sys.path.append("/Users/kuramochiosuke/.pyenv/versions/3.10.4/lib/python3.10/site-packages")
from dotenv import find_dotenv, load_dotenv
import requests
import schedule
import time
import openpyxl

def job():
    try:
        # .envファイルを探して読み込み
        env_file = find_dotenv()
        load_dotenv(env_file)  

        CONSUMER_KEY = os.environ.get('CONSUMER_KEY')
        CONSUMER_SECRET = os.environ.get('CONSUMER_SECRET')
        ACCESS_KEY = os.environ.get('ACCESS_KEY')
        ACCESS_KEY_SECRET = os.environ.get('ACCESS_KEY_SECRET')

        # Twitterの認証
        twitter = OAuth1Session(CONSUMER_KEY, CONSUMER_SECRET, ACCESS_KEY, ACCESS_KEY_SECRET)
        print(twitter)
        #エンドポイント
        url_text = 'https://api.twitter.com/1.1/statuses/update.json'
        url_media = "https://upload.twitter.com/1.1/media/upload.json"
        print("投稿準備完了")
        #nbaのページの指定
        URL = 'https://basketballking.jp/news/category/world?cx_cat=head'
        # リクエストヘッダの指定
        headers = {"User-Agent": "hoge"}
        response = requests.get(URL,  headers=headers)
        r_text=response.text
        soup = BeautifulSoup(r_text, 'html.parser')
        soup_a=soup.find_all("div",attrs={"class","container"})[1]
        soup_article=soup_a.find_all("div",attrs={"class","news-category-list__vertical"})[0]
        soup_title=soup_article.find_all("div",attrs={"class","news-category-list__vertical__title"})[0].text.replace("\n","").replace("\t","")
        soup_url=soup_article.find_all("a")[0]['href']
        soup_img=soup_article.find("img")["src"]
        response = requests.get(soup_img)
        image = response.content
        files = {"media" : image}
        req_media = twitter.post(url_media, files = files)
        media_id = json.loads(req_media.text)['media_id']
        print('画像の取得完了')
        params = {'status':"バスケ情報!!!\n\n{}\n\n記事↓↓↓\n{}".format(soup_title,soup_url),'media_ids':[media_id]}
        wb = openpyxl.load_workbook('nba.xlsx')
        time.sleep(1)
        ws = wb["Sheet1"]
        for i in range(wb['Sheet1'].max_row):
            if ws.cell(row=i+1,column=1).value==params["status"]:
                print("投稿済みです")
                break    
            elif i==wb['Sheet1'].max_row-1:   
                ws.cell(row=wb['Sheet1'].max_row+1,column=1).value = params["status"]
                wb.save('nba.xlsx')
                print("保存しました")
                twitter.post(url_text, params = params)
                print("投稿しました") 
        print("処理終了しました")
        print("")
    except IndexError:
        print("INDEX エラーです")       
        print("")
    
    except FileNotFoundError:
        print("NOT FILE エラーです")       
        print("")

    except KeyError:
        print("KeyError エラーです")       
        print("")

def main():
    schedule.every(3).minutes.do(job)
    while True:
        schedule.run_pending()
        time.sleep(1)

if __name__ == '__main__':
    main()
